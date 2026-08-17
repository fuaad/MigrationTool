"""reporting.py - Export the tracking state to an Excel audit report."""

from __future__ import annotations

import json
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill


def export_report(tracker, out_path):
    wb = openpyxl.Workbook()

    # --- Projects sheet ---
    ws = wb.active
    ws.title = "Projects"
    headers = ["Project Number", "Project Title", "Program", "Business Line",
               "Master ID", "Project ID", "Status", "Error", "Updated"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="1F4E5F")
        c.font = Font(bold=True, color="FFFFFF")
    for p in tracker.all_projects():
        ws.append([p["project_number"], p["project_title"], p["program"],
                   p["business_line"], p["master_id"], p["project_id"],
                   p["status"], p["error"], p["updated_at"]])

    # --- Documents sheet ---
    wd = wb.create_sheet("Documents")
    dheaders = ["Project Number", "File Name", "Folder Path", "Match", "Revision Number",
                "Revision Stage", "Document ID", "Rev Status ID", "Status",
                "Error", "Source Path", "Updated"]
    wd.append(dheaders)
    for c in wd[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E5F")
    for p in tracker.all_projects():
        for d in tracker.documents_for(p["project_number"]):
            wd.append([d["project_number"], d["file_name"], d["folder_path"],
                       d["match_status"], d["revision_number"], d["revision_stage"],
                       d["document_id"], d["rev_status_id"], d["status"], d["error"],
                       d["source_path"], d["updated_at"]])

    # --- Summary sheet ---
    wss = wb.create_sheet("Summary", 0)
    proj_sum, doc_sum, match_sum = tracker.summary()
    wss.append(["xENG Migration - Status Summary"])
    wss["A1"].font = Font(bold=True, size=14)
    wss.append(["Generated", datetime.datetime.now().isoformat(timespec="seconds")])
    wss.append([])
    wss.append(["Projects by status"])
    wss["A4"].font = Font(bold=True)
    for k, v in sorted(proj_sum.items()):
        wss.append([k, v])
    wss.append([])
    row = wss.max_row + 1
    wss.append(["Documents by status"])
    wss.cell(row=wss.max_row, column=1).font = Font(bold=True)
    for k, v in sorted(doc_sum.items()):
        wss.append([k, v])
    wss.append([])
    wss.append(["Documents by match status"])
    wss.cell(row=wss.max_row, column=1).font = Font(bold=True)
    for k, v in sorted(match_sum.items()):
        wss.append([k, v])

    # autosize-ish
    for sheet in (wss, ws, wd):
        for col in sheet.columns:
            width = max((len(str(c.value)) if c.value else 0) for c in col) + 2
            sheet.column_dimensions[col[0].column_letter].width = min(width, 60)

    wb.save(out_path)
    return out_path

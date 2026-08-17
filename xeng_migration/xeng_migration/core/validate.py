r"""validate.py - Preflight validation of the register against disk and data.

Runs BEFORE any ECM calls. For every document row it classifies the exact
problem (if any) so you get an actionable report instead of a single count.

Checks per document:
  OK                    - source file exists and is readable
  ROOT_MISSING          - the configured source.root does not exist/instrument
  PROJECT_FOLDER_MISSING- <root>\<project> folder not found
  FOLDER_MISSING        - the Folder Path folder not found on disk
  FILE_MISSING          - folder exists but the file is not in it
  PERMISSION_DENIED     - path exists but cannot be read
  PATH_TOO_LONG         - exceeds Windows MAX_PATH (>259 chars)
  ORPHAN_PROJECT        - register row's project not in the project list
  NO_SOURCE_PATH        - could not build a source path (missing columns)

Also reports (project-level / folder-level):
  UNREGISTERED_FILES    - files present on disk but not in the register
  DUPLICATE_ROWS        - same project/folder/file listed more than once
"""

from __future__ import annotations

import os


WIN_MAXPATH = 259


def _classify_path(root, source_path, project_number):
    # Root itself
    if not root or not os.path.isdir(root):
        return "ROOT_MISSING", f"source.root not found: {root}"

    if not source_path:
        return "NO_SOURCE_PATH", "could not build source path (missing columns)"

    if len(source_path) > WIN_MAXPATH:
        return "PATH_TOO_LONG", f"{len(source_path)} chars (> {WIN_MAXPATH})"

    # Walk down to give a precise 'where it broke' answer
    project_dir = os.path.join(root, str(project_number))
    if not os.path.isdir(project_dir):
        return "PROJECT_FOLDER_MISSING", project_dir

    folder = os.path.dirname(source_path)
    if not os.path.isdir(folder):
        return "FOLDER_MISSING", folder

    if not os.path.exists(source_path):
        return "FILE_MISSING", source_path

    # Exists - can we read it?
    try:
        with open(source_path, "rb") as f:
            f.read(1)
    except PermissionError:
        return "PERMISSION_DENIED", source_path
    except OSError as e:
        return "READ_ERROR", f"{source_path} ({e})"

    return "OK", source_path


def validate(tracker, cfg):
    root = cfg["source"]["root"]
    known_projects = {p["project_number"] for p in tracker.all_projects()}

    results = []           # (project, folder, file, status, detail)
    seen = {}              # (proj,folder,file) -> count for dup detection
    per_status = {}

    for p in tracker.all_projects():
        for d in tracker.documents_for(p["project_number"]):
            key = (d["project_number"], d["folder_path"], d["file_name"])
            seen[key] = seen.get(key, 0) + 1

            if d["project_number"] not in known_projects:
                status, detail = "ORPHAN_PROJECT", d["project_number"]
            else:
                status, detail = _classify_path(
                    root, d["source_path"], d["project_number"])

            per_status[status] = per_status.get(status, 0) + 1
            results.append((d["project_number"], d["folder_path"],
                            d["file_name"], status, detail))

    # duplicates
    dups = [(k, c) for k, c in seen.items() if c > 1]

    # unregistered files on disk (per project folder that exists)
    unregistered = _find_unregistered(tracker, root, known_projects,
                                       cfg["source"].get("project_folder_is_number", True))

    return {
        "results": results,
        "per_status": per_status,
        "duplicates": dups,
        "unregistered": unregistered,
    }


def _find_unregistered(tracker, root, known_projects, project_folder_is_number):
    """Scan-driven model: the scan IS the inventory, so there are no
    'unregistered' files. Kept for report compatibility."""
    return []
    # (legacy register-driven logic below, unreachable)
    if not root or not os.path.isdir(root):
        return []
    registered = set()
    for p in tracker.all_projects():
        for d in tracker.documents_for(p["project_number"]):
            if d["source_path"]:
                registered.add(os.path.normcase(os.path.abspath(d["source_path"])))

    unregistered = []
    for proj in known_projects:
        base = os.path.join(root, str(proj)) if project_folder_is_number else root
        if not os.path.isdir(base):
            continue
        for dirpath, _dirs, files in os.walk(base):
            for fn in files:
                full = os.path.normcase(os.path.abspath(os.path.join(dirpath, fn)))
                if full not in registered:
                    unregistered.append(os.path.join(dirpath, fn))
    return unregistered


def write_report(validation, out_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Validation"
    ws.append(["Project", "Folder Path", "File Name", "Status", "Detail"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="7A1F1F")
    red = PatternFill("solid", fgColor="F8D7DA")
    green = PatternFill("solid", fgColor="D4EDDA")
    for proj, folder, fname, status, detail in validation["results"]:
        ws.append([proj, folder, fname, status, detail])
        fill = green if status == "OK" else red
        ws.cell(row=ws.max_row, column=4).fill = fill

    # summary
    wss = wb.create_sheet("Summary", 0)
    wss.append(["Preflight Validation Summary"])
    wss["A1"].font = Font(bold=True, size=14)
    wss.append([])
    wss.append(["Status", "Count"])
    wss["A3"].font = Font(bold=True)
    for k, v in sorted(validation["per_status"].items()):
        wss.append([k, v])
    wss.append([])
    wss.append(["Duplicate rows", len(validation["duplicates"])])
    wss.append(["Unregistered files on disk", len(validation["unregistered"])])

    if validation["duplicates"]:
        wd = wb.create_sheet("Duplicates")
        wd.append(["Project", "Folder Path", "File Name", "Count"])
        for (proj, folder, fname), c in validation["duplicates"]:
            wd.append([proj, folder, fname, c])

    if validation["unregistered"]:
        wu = wb.create_sheet("Unregistered")
        wu.append(["File on disk not in register"])
        for f in validation["unregistered"]:
            wu.append([f])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            width = max((len(str(c.value)) if c.value else 0) for c in col) + 2
            sheet.column_dimensions[col[0].column_letter].width = min(width, 80)

    wb.save(out_path)
    return out_path

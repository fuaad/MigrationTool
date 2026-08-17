"""match.py - Match SLFE records to scanned documents by file name.

Rules (never guess):
  - A record matches by exact file name (case-insensitive) within its project.
  - Exactly ONE candidate  -> MATCHED: record's metadata + revision copied
    onto the document row.
  - MORE than one candidate -> AMBIGUOUS: nothing applied; all candidates
    flagged AMBIGUOUS and listed in the review report.
  - ZERO candidates        -> record flagged UNMATCHED (file not on drive).
Documents never matched by any record stay UNMATCHED (structure-only
migration).
"""

from __future__ import annotations


def match(tracker, project_number=None):
    projects = ([tracker.get_project(project_number)] if project_number
                else tracker.all_projects())
    projects = [p for p in projects if p]

    stats = {"matched": 0, "ambiguous_records": 0, "unmatched_records": 0}
    review = []      # (project, record file name, [candidate folder paths])
    missing = []     # (project, record file name) records with no file

    for p in projects:
        pnum = p["project_number"]
        for rec in tracker.records_for(pnum):
            candidates = tracker.documents_by_name(pnum, rec["file_name"])

            if len(candidates) == 1:
                d = candidates[0]
                tracker.set_document(
                    d["id"],
                    match_status="MATCHED",
                    record_id=rec["id"],
                    revision_number=rec["revision_number"],
                    revision_stage=rec["revision_stage"],
                    meta_json=rec["meta_json"],
                )
                tracker.set_record(rec["id"], match_status="MATCHED",
                                   matched_doc_id=d["id"])
                stats["matched"] += 1

            elif len(candidates) > 1:
                for d in candidates:
                    tracker.set_document(d["id"], match_status="AMBIGUOUS")
                tracker.set_record(rec["id"], match_status="AMBIGUOUS")
                review.append((pnum, rec["file_name"],
                               [d["folder_path"] for d in candidates]))
                stats["ambiguous_records"] += 1

            else:
                tracker.set_record(rec["id"], match_status="UNMATCHED")
                missing.append((pnum, rec["file_name"]))
                stats["unmatched_records"] += 1

        tracker.log("INFO", pnum, f"Match complete: {stats}")

    return {"stats": stats, "review": review, "missing": missing}


def write_match_report(result, tracker, out_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()

    wss = wb.active
    wss.title = "Summary"
    wss.append(["Record Matching Summary"])
    wss["A1"].font = Font(bold=True, size=14)
    wss.append([])
    s = result["stats"]
    wss.append(["Records matched (metadata will be applied)", s["matched"]])
    wss.append(["Records ambiguous (needs review)", s["ambiguous_records"]])
    wss.append(["Records with no file on drive", s["unmatched_records"]])
    _proj, _doc, match_sum = tracker.summary()
    wss.append([])
    wss.append(["Documents on drive by match status"])
    wss.cell(row=wss.max_row, column=1).font = Font(bold=True)
    for k, v in sorted(match_sum.items()):
        wss.append([k, v])

    if result["review"]:
        wr = wb.create_sheet("Needs Review")
        wr.append(["Project", "Record File Name", "Candidate Locations (choose one)"])
        for c in wr[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="B45309")
        for pnum, fname, folders in result["review"]:
            wr.append([pnum, fname, " | ".join(f or "(project root)" for f in folders)])

    if result["missing"]:
        wm = wb.create_sheet("Records Without File")
        wm.append(["Project", "Record File Name (not found on drive)"])
        for c in wm[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="7A1F1F")
        for pnum, fname in result["missing"]:
            wm.append([pnum, fname])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            width = max((len(str(c.value)) if c.value else 0) for c in col) + 2
            sheet.column_dimensions[col[0].column_letter].width = min(width, 80)

    wb.save(out_path)
    return out_path

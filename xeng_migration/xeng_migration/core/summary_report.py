"""summary_report.py - Post-migration counts of what actually landed in ECM.

Produces an Excel workbook with:
  * Overview      - project totals (folders / documents) as counted in ECM,
                    alongside what the tracker expected from the drive scan.
  * By Top Folder - per first-level folder (e.g. '4000 ENG'): number of
                    sub-folders and number of documents beneath it (recursive).
  * Failures      - any documents not COMPLETE, with their error.

Counts come from the tracker's record of what was created (fast, no ECM
crawl). Use --verify to additionally count live in ECM via REST.
"""

from __future__ import annotations

import json


def _gb(num_bytes):
    """Bytes -> GB, rounded to 2 decimals (0 stays 0)."""
    try:
        return round(int(num_bytes or 0) / (1024 ** 3), 2)
    except (TypeError, ValueError):
        return 0


def _top(folder_path):
    if not folder_path:
        return "(project root)"
    return folder_path.replace("/", "\\").split("\\", 1)[0]


def _timing(tracker, pnum):
    """Derive elapsed time from tracker timestamps and the completion log."""
    rows = tracker.conn.execute(
        "SELECT MIN(created_at), MAX(updated_at) FROM documents "
        "WHERE project_number=?", (pnum,)).fetchone()
    first, last = (rows[0], rows[1]) if rows else (None, None)
    log = tracker.conn.execute(
        "SELECT message FROM run_log WHERE scope=? AND message LIKE "
        "'Project completed in%' ORDER BY id DESC LIMIT 1", (pnum,)).fetchone()
    return {"first_activity": first, "last_activity": last,
            "completion_note": log[0] if log else None}


def build(tracker, project_number=None):
    projects = ([tracker.get_project(project_number)] if project_number
                else tracker.all_projects())
    projects = [p for p in projects if p]

    out = []
    for p in projects:
        pnum = p["project_number"]
        docs = tracker.documents_for(pnum)

        migrated = [d for d in docs if d["status"] == "COMPLETE"]
        failed = [d for d in docs if d["status"] not in ("COMPLETE",)]

        # distinct folders that actually received documents
        folders_used = {d["target_folder_id"] for d in migrated
                        if d["target_folder_id"]}
        # distinct folder PATHS (includes intermediate levels)
        all_paths = set()
        for d in docs:
            fp = (d["folder_path"] or "").replace("/", "\\")
            parts = [s for s in fp.split("\\") if s.strip()]
            for i in range(1, len(parts) + 1):
                all_paths.add("\\".join(parts[:i]))

        def _size(d):
            try:
                return int(d["file_size"] or 0)
            except (KeyError, IndexError, TypeError, ValueError):
                return 0

        by_top = {}
        for d in docs:
            t = _top(d["folder_path"])
            e = by_top.setdefault(t, {"docs": 0, "done": 0, "failed": 0,
                                      "paths": set(), "bytes": 0,
                                      "bytes_done": 0})
            e["docs"] += 1
            sz = _size(d)
            e["bytes"] += sz
            if d["status"] == "COMPLETE":
                e["done"] += 1
                e["bytes_done"] += sz
            else:
                e["failed"] += 1
            fp = (d["folder_path"] or "").replace("/", "\\")
            parts = [s for s in fp.split("\\") if s.strip()]
            for i in range(1, len(parts) + 1):
                e["paths"].add("\\".join(parts[:i]))

        tinfo = _timing(tracker, pnum)
        try:
            mig_start = p["migration_start"]
            mig_end = p["migration_end"]
        except (KeyError, IndexError):
            mig_start = mig_end = None
        out.append({
            "migration_start": mig_start,
            "migration_end": mig_end,
            "first_activity": tinfo["first_activity"],
            "last_activity": tinfo["last_activity"],
            "completion_note": tinfo["completion_note"],
            "project_number": pnum,
            "project_title": p["project_title"],
            "master_id": p["master_id"],
            "project_id": p["project_id"],
            "status": p["status"],
            "total_documents": len(docs),
            "migrated": len(migrated),
            "failed": len(failed),
            "folders_total": len(all_paths),
            "folders_with_documents": len(folders_used),
            "matched": sum(1 for d in docs if d["match_status"] == "MATCHED"),
            "ambiguous": sum(1 for d in docs if d["match_status"] == "AMBIGUOUS"),
            "unmatched": sum(1 for d in docs if d["match_status"] == "UNMATCHED"),
            "total_bytes": sum(_size(d) for d in docs),
            "migrated_bytes": sum(_size(d) for d in migrated),
            "by_top": {k: {"docs": v["docs"], "done": v["done"],
                           "failed": v["failed"], "folders": len(v["paths"]),
                           "bytes": v["bytes"], "bytes_done": v["bytes_done"]}
                       for k, v in sorted(by_top.items())},
            "failures": [(d["file_name"], d["folder_path"], d["status"],
                          (d["error"] or "")[:300]) for d in failed],
        })
    return out


def write(summaries, out_path):
    """Write the summary workbook in the agreed client layout:

        Migration Start Date Time | <value>
        Migration End Date Time   | <value>
        [ Project | First-level folder | Sub-folders | Documents |
          Migrated | Failed ]
        ... one row per first-level folder ...
        Total row
        (blank)
        Failures details
        [ File Name | Folder Path | Status | Error ]
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HEAD_FILL = PatternFill("solid", fgColor="1F4E5F")
    HEAD_FONT = Font(bold=True, color="FFFFFF")
    LABEL_FILL = PatternFill("solid", fgColor="1F4E5F")
    thin = Side(style="thin", color="9BB7C4")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Migration Summary"

    # column widths follow the layout: A project, B folder name, C-F counts
    for col, width in (("A", 24), ("B", 58), ("C", 13), ("D", 13),
                       ("E", 11), ("F", 9), ("G", 12), ("H", 14)):
        ws.column_dimensions[col].width = width

    r = 1
    for s in summaries:
        # --- header: start / end date time ---
        for label, value in (("Migration Start Date Time",
                              s.get("migration_start") or ""),
                             ("Migration End Date Time",
                              s.get("migration_end") or "")):
            c = ws.cell(row=r, column=1, value=label)
            c.font = HEAD_FONT
            c.fill = LABEL_FILL
            c.border = BORDER
            v = ws.cell(row=r, column=2, value=value)
            v.border = BORDER
            ws.merge_cells(start_row=r, start_column=2, end_row=r,
                           end_column=8)
            r += 1

        # --- table header ---
        headers = ["Project", "First-level folder", "Sub-folders",
                   "Documents", "Migrated", "Failed", "Size (GB)",
                   "Migrated (GB)"]
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=r, column=i, value=h)
            c.font = HEAD_FONT
            c.fill = HEAD_FILL
            c.alignment = CENTER
            c.border = BORDER
        r += 1

        # --- one row per first-level folder ---
        tot_folders = tot_docs = tot_done = tot_failed = 0
        tot_bytes = tot_bytes_done = 0
        for top, v in s["by_top"].items():
            ws.cell(row=r, column=1, value=s["project_number"]).border = BORDER
            ws.cell(row=r, column=2, value=top).border = BORDER
            for i, val in enumerate((v["folders"], v["docs"], v["done"],
                                     v["failed"], _gb(v["bytes"]),
                                     _gb(v["bytes_done"])), start=3):
                c = ws.cell(row=r, column=i, value=val)
                c.border = BORDER
                c.alignment = CENTER
                if i >= 7:
                    c.number_format = "0.00"
            tot_folders += v["folders"]
            tot_docs += v["docs"]
            tot_done += v["done"]
            tot_failed += v["failed"]
            tot_bytes += v["bytes"]
            tot_bytes_done += v["bytes_done"]
            r += 1

        # --- total row ---
        tc = ws.cell(row=r, column=2, value="Total")
        tc.font = HEAD_FONT
        tc.fill = HEAD_FILL
        tc.alignment = Alignment(horizontal="right")
        tc.border = BORDER
        ws.cell(row=r, column=1).fill = HEAD_FILL
        ws.cell(row=r, column=1).border = BORDER
        for i, val in enumerate((tot_folders, tot_docs, tot_done,
                                 tot_failed, _gb(tot_bytes),
                                 _gb(tot_bytes_done)), start=3):
            c = ws.cell(row=r, column=i, value=val)
            c.border = BORDER
            c.alignment = CENTER
            c.font = Font(bold=True)
            if i >= 7:
                c.number_format = "0.00"
        r += 2

        # --- failures section ---
        fh = ws.cell(row=r, column=1, value="Failures details")
        fh.font = HEAD_FONT
        fh.fill = HEAD_FILL
        fh.alignment = CENTER
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER
        r += 1

        for i, h in enumerate(["File Name", "Folder Path", "Status", "Error"],
                              start=1):
            c = ws.cell(row=r, column=i, value=h)
            c.font = HEAD_FONT
            c.fill = HEAD_FILL
            c.alignment = CENTER
            c.border = BORDER
        r += 1

        if s["failures"]:
            for fn, fp, st, err in s["failures"]:
                ws.cell(row=r, column=1, value=fn).border = BORDER
                ws.cell(row=r, column=2, value=fp).border = BORDER
                ws.cell(row=r, column=3, value=st).border = BORDER
                ws.cell(row=r, column=4, value=err).border = BORDER
                r += 1
        else:
            ws.cell(row=r, column=1, value="(none)").border = BORDER
            for col in range(2, 5):
                ws.cell(row=r, column=col).border = BORDER
            r += 1

        r += 2   # gap before the next project block

    wb.save(out_path)
    return out_path


def verify_live(otcs, project_id, top_folder_ids=None):
    """Optional: count nodes live in ECM under the project (recursive).

    Returns {'folders': n, 'documents': n}. Slower - walks the tree via REST.
    """
    folders = 0
    documents = 0
    stack = [project_id]
    while stack:
        nid = stack.pop()
        children = otcs.list_children(nid)
        for cid, ctype in children:
            if ctype == 0:
                folders += 1
                stack.append(cid)
            else:
                documents += 1
    return {"folders": folders, "documents": documents}
